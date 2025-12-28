# save as excel_graph_openpyxl_xlwings.py
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import Scaling, ChartLines
import math
import os

# ===== 保存先（デスクトップ） =====
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
filepath = os.path.join(desktop, "test.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# ===== データ作成 =====
ws["H2"] = "グラフ名"
ws["H4"] = "系列名"

n = 19
angles = [round(-90 + i * (180 / (n - 1)), 6) for i in range(n)]
values = [round(math.cos(math.radians(a)), 6) for a in angles]

angle_row, angle_col = 3, 9  # I3
value_row, value_col = 4, 9  # I4

# データ書き込み
for i in range(n):
    ws.cell(row=angle_row, column=angle_col + i, value=angles[i])
    ws.cell(row=value_row, column=value_col + i, value=values[i])

# ===== グラフ作成 =====
chart = ScatterChart()
chart.title = ws["H2"].value

chart.x_axis.title = "角度 (deg.)"
chart.y_axis.title = "cos"

chart.x_axis.scaling = Scaling(min=-90, max=90)
chart.x_axis.majorUnit = 15
chart.y_axis.scaling = Scaling(min=-1, max=1)
chart.y_axis.majorUnit = 0.5

chart.x_axis.majorGridlines = ChartLines()
chart.y_axis.majorGridlines = ChartLines()

# 凡例を消す
chart.legend = None

# ===== データ範囲 =====
min_col = angle_col
max_col = angle_col + (n - 1)

xvalues = Reference(ws, min_col=min_col, max_col=max_col, min_row=angle_row)
yvalues = Reference(ws, min_col=min_col, max_col=max_col, min_row=value_row)
series = Series(yvalues, xvalues, title=None)

# ---- 色を青に統一 ----
blue = "4472C4"
series.graphicalProperties.line.solidFill = blue
series.marker.symbol = "circle"
series.marker.size = 7
series.marker.graphicalProperties.line.solidFill = blue
series.marker.graphicalProperties.solidFill = blue

chart.series.append(series)

chart.width = 11.5
chart.height = 5.9

# ===== A1 に貼り付ける =====
ws.add_chart(chart, "A1")

wb.save(filepath)
print("Saved:", filepath)

# modify.py
import xlwings as xw
from xlwings.constants import AxisType
from win32com.client import constants

app = xw.App(visible=False)         # Excel を表示しない
wb = app.books.open(filepath)

def RGB(r, g, b):
    return r + (g << 8) + (b << 16)

ws = wb.sheets[0]               # 1つ目のシートを指定

# シート内のグラフ名を全部表示して確認
chart_count = ws.api.ChartObjects().Count
print("Chart count:", chart_count)

for i in range(1, chart_count + 1):
    obj = ws.api.ChartObjects(i)
    print(i, obj.Name)

# 1番目のグラフを取得（必要なら Name 指定も可）
obj = ws.api.ChartObjects(1)
chart = obj.Chart
ch=chart

# 横軸のオプション
x_axis = ch.Axes(AxisType.xlCategory)
x_axis.MinimumScale = -90  # 最小値
x_axis.MaximumScale = 90   # 最大値
x_axis.MajorUnit = 15      # 目盛間隔
x_axis.CrossesAt = -90     # 交差位置(縦軸との交点)
x_axis.TickLabels.NumberFormatLocal = "0"   # 小数0桁まで表示

# 横軸のタイトル
x_axis.HasTitle = True
if x_axis.HasTitle:
    x_axis.AxisTitle.Text = "角度 (deg.)"

# 縦軸のオプション
y_axis = ch.Axes(AxisType.xlValue)
y_axis.TickLabels.NumberFormatLocal = "0.0" # 小数1桁まで表示

# 縦軸のタイトル
y_axis.HasTitle = False
if y_axis.HasTitle:
    y_axis.AxisTitle.Text = ""

y_axis.MinimumScale = 0  # 最小値

# Excel 2021 以降の標準スタイルを指定する ----------------------------------
# グラフタイトルの文字色をRGB(89,89,89)とし、フォントサイズを14にする
if ch.HasTitle:
    ch.ChartTitle.Format.TextFrame2.TextRange.Font.Bold = False
    ch.ChartTitle.Format.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = RGB(89,89,89)
    ch.ChartTitle.Format.TextFrame2.TextRange.Font.Size = 14

# グリッド線の設定（薄いグレー)
x_axis.HasMajorGridlines = True
x_axis.MajorGridlines.Format.Line.ForeColor.RGB = RGB(217, 217, 217)
x_axis.MajorGridlines.Format.Line.Weight = 0.75
x_axis.Format.Line.ForeColor.RGB = RGB(191, 191, 191)
x_axis.MajorTickMark = constants.xlTickMarkNone # 目盛の内向き/外向きなし
y_axis.HasMajorGridlines = True
y_axis.MajorGridlines.Format.Line.ForeColor.RGB = RGB(217, 217, 217)
y_axis.MajorGridlines.Format.Line.Weight = 0.75
y_axis.Format.Line.ForeColor.RGB = RGB(191, 191, 191)
y_axis.MajorTickMark = constants.xlTickMarkNone # 目盛の内向き/外向きなし

for ax in ch.Axes():    
    # 軸の設定
    ax.TickLabels.Font.Color = RGB(89,89,89)
    ax.TickLabels.Font.Size = 9
    ax.TickLabels.Font.Name = "Aptos Narrow 本文"
    # 軸タイトルがあるとき、軸タイトルを設定する
    if ax.HasTitle:
        ax.AxisTitle.Format.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = RGB(89,89,89)
        ax.AxisTitle.Format.TextFrame2.TextRange.Font.Bold = False
        ax.AxisTitle.Format.TextFrame2.TextRange.Font.Size = 10

# 外枠の設定
ch.ChartArea.Format.Line.ForeColor.RGB = RGB(217,217,217) # 薄いグレー
ch.ChartArea.Format.Line.Weight = 0.75                    # 枠線の太さ(pt)
# ----------------------------------------------------------------------

# 凡例なし
ch.HasLegend = False
    
# プロットエリアの調整
p = ch.PlotArea
p.InsideLeft   = p.InsideLeft
p.InsideTop    = p.InsideTop
p.InsideWidth  = p.InsideWidth
p.InsideHeight = p.InsideHeight+15



# 軸タイトルと目盛りの数値の色を黒に変更する
for ax in ch.Axes():
    ax.TickLabels.Font.Color = RGB(0,0,0)
    if ax.HasTitle:
        ax.AxisTitle.Format.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = RGB(0,0,0)
        
# グラフ外枠を黒に変更
ch.ChartArea.Format.Line.ForeColor.RGB = RGB(0,0,0)
wb.save()
wb.close()
app.quit()
